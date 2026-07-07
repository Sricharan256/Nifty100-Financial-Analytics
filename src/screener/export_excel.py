"""
Sprint 3 - Day 17

Generate screener_output.xlsx
"""

import os
import sqlite3
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DB_PATH = "db/nifty100.db"

OUTPUT = "output/screener_output.xlsx"


# -------------------------------------------------------
# Load Latest Financial Ratios
# -------------------------------------------------------

def load_latest():

    conn = sqlite3.connect(DB_PATH)

    df = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    conn.close()

    df = (
        df.sort_values("year")
        .groupby("company_id")
        .last()
        .reset_index()
    )

    return df


# -------------------------------------------------------
# Presets
# -------------------------------------------------------

def quality(df):

    return df[
        (df["return_on_equity_pct"] >= 20)
        &
        (df["debt_to_equity"] <= 0.5)
        &
        (df["free_cash_flow_cr"] > 0)
    ]


def value(df):

    return df[
        (df["return_on_equity_pct"] >= 15)
        &
        (df["debt_to_equity"] <= 1)
    ]


def growth(df):

    return df[
        (df["return_on_equity_pct"] >= 20)
    ]


def dividend(df):

    return df[
        (df["dividend_payout_ratio_pct"] <= 80)
        &
        (df["free_cash_flow_cr"] > 0)
    ]


def debt_free(df):

    return df[
        df["debt_to_equity"] == 0
    ]


def turnaround(df):

    return df[
        (df["free_cash_flow_cr"] > 0)
        &
        (df["return_on_equity_pct"] >= 10)
    ]


# -------------------------------------------------------
# Export Excel
# -------------------------------------------------------

def export():

    os.makedirs("output", exist_ok=True)

    df = load_latest()

    presets = {

        "Quality Compounder": quality(df),

        "Value Pick": value(df),

        "Growth Accelerator": growth(df),

        "Dividend Champion": dividend(df),

        "Debt Free Blue Chip": debt_free(df),

        "Turnaround Watch": turnaround(df)

    }

    with pd.ExcelWriter(
        OUTPUT,
        engine="openpyxl"
    ) as writer:

        for name, data in presets.items():

            data = data.sort_values(
                "composite_quality_score",
                ascending=False
            )

            data.to_excel(
                writer,
                sheet_name=name,
                index=False
            )

    colour_excel()


# -------------------------------------------------------
# Colour Formatting
# -------------------------------------------------------

def colour_excel():

    wb = load_workbook(OUTPUT)

    green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        headers = {}

        for cell in ws[1]:

            headers[cell.column] = cell.value

        for row in range(2, ws.max_row + 1):

            for col in headers:

                header = headers[col]

                value = ws.cell(row, col).value

                if value is None:
                    continue

                try:

                    if header == "return_on_equity_pct":

                        if value >= 15:
                            ws.cell(row, col).fill = green
                        else:
                            ws.cell(row, col).fill = red

                    elif header == "debt_to_equity":

                        if value <= 1:
                            ws.cell(row, col).fill = green
                        else:
                            ws.cell(row, col).fill = red

                    elif header == "free_cash_flow_cr":

                        if value > 0:
                            ws.cell(row, col).fill = green
                        else:
                            ws.cell(row, col).fill = red

                    elif header == "composite_quality_score":

                        if value >= 20:
                            ws.cell(row, col).fill = green
                        else:
                            ws.cell(row, col).fill = red

                except:
                    pass

    wb.save(OUTPUT)


# -------------------------------------------------------
# Main
# -------------------------------------------------------

if __name__ == "__main__":

    print("=" * 60)
    print("SPRINT 3 DAY 17")
    print("=" * 60)

    export()

    print()

    print("Excel Report Generated Successfully")

    print(f"Location : {OUTPUT}")