"""
peer_comparison_report.py

Sprint 3 - Day 20

Peer Comparison Excel Report
"""

import sqlite3
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

DB_PATH = "db/nifty100.db"

OUTPUT = "output/peer_comparison.xlsx"


# ---------------------------------------------------------
# Load Companies
# ---------------------------------------------------------

def load_companies():

    conn = sqlite3.connect(DB_PATH)

    companies = pd.read_sql(
        """
        SELECT
            id AS company_id,
            company_name
        FROM companies
        """,
        conn
    )

    conn.close()

    return companies


# ---------------------------------------------------------
# Load Financial Ratios
# ---------------------------------------------------------

def load_financial_ratios():

    conn = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql(
        """
        SELECT *
        FROM financial_ratios
        """,
        conn
    )

    conn.close()

    ratios = (
        ratios
        .sort_values("year")
        .groupby("company_id", as_index=False)
        .last()
    )

    return ratios


# ---------------------------------------------------------
# Load Peer Percentiles
# ---------------------------------------------------------

def load_peer_percentiles():

    conn = sqlite3.connect(DB_PATH)

    peer = pd.read_sql(
        """
        SELECT *
        FROM peer_percentiles
        """,
        conn
    )

    conn.close()

    return peer


# ---------------------------------------------------------
# Merge Data
# ---------------------------------------------------------

def prepare_data():

    companies = load_companies()

    ratios = load_financial_ratios()

    peer = load_peer_percentiles()

    df = pd.merge(
        ratios,
        companies,
        on="company_id",
        how="left"
    )

    df = pd.merge(
        df,
        peer[
            [
                "company_id",
                "peer_group_name"
            ]
        ].drop_duplicates(),
        on="company_id",
        how="left"
    )

    df["peer_group_name"] = df[
        "peer_group_name"
    ].fillna("No Peer Group")

    return df


# ---------------------------------------------------------
# Peer Groups
# ---------------------------------------------------------

def get_peer_groups(df):

    return sorted(
        df["peer_group_name"]
        .dropna()
        .unique()
    )
# ---------------------------------------------------------
# Report Columns
# ---------------------------------------------------------

REPORT_COLUMNS = [

    "company_id",
    "company_name",

    "return_on_equity_pct",
    "return_on_capital_employed_pct",

    "net_profit_margin_pct",
    "operating_profit_margin_pct",

    "debt_to_equity",
    "interest_coverage_ratio",

    "asset_turnover_ratio",

    "free_cash_flow_cr",
    "cfo_quality_score",

    "fcf_conversion_pct",

    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    "composite_quality_score",

    "peer_group_name"
]


# ---------------------------------------------------------
# Prepare Peer Sheet
# ---------------------------------------------------------

def prepare_peer_sheet(df, peer_group):

    sheet = df[
        df["peer_group_name"] == peer_group
    ].copy()

    columns = []

    for col in REPORT_COLUMNS:

        if col in sheet.columns:

            columns.append(col)

    sheet = sheet[columns]

    sheet = sheet.sort_values(

        by="composite_quality_score",

        ascending=False

    )

    return sheet


# ---------------------------------------------------------
# Create Excel Workbook
# ---------------------------------------------------------

def create_workbook():

    os.makedirs(
        "output",
        exist_ok=True
    )

    df = prepare_data()

    peer_groups = get_peer_groups(df)

    writer = pd.ExcelWriter(

        OUTPUT,

        engine="openpyxl"

    )

    for group in peer_groups:

        sheet = prepare_peer_sheet(

            df,

            group

        )

        sheet_name = group[:31]

        sheet.to_excel(

            writer,

            sheet_name=sheet_name,

            index=False

        )

    return writer, peer_groups
# ---------------------------------------------------------
# Save Workbook
# ---------------------------------------------------------

def save_workbook():

    writer, peer_groups = create_workbook()

    writer.close()

    print("=" * 60)
    print("SPRINT 3 - DAY 20")
    print("PEER COMPARISON EXCEL REPORT")
    print("=" * 60)

    print(f"\nPeer Groups Exported : {len(peer_groups)}")

    print(f"\nExcel Report Generated : {OUTPUT}")

    print("\nSheets Created:\n")

    for group in peer_groups:

        print(f"✔ {group}")

    print("\nWorkbook Created Successfully.")
    format_workbook()
    finalize_workbook()

def preview():

    df = prepare_data()

    print("\nPreview\n")

    print(
        df[
            [
                "company_id",
                "company_name",
                "peer_group_name"
            ]
        ].head(20)
    )
# ---------------------------------------------------------
# Format Workbook
# ---------------------------------------------------------

def format_workbook():

    wb = load_workbook(OUTPUT)

    green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    yellow = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="F4CCCC",
        end_color="F4CCCC",
        fill_type="solid"
    )

    for ws in wb.worksheets:

        # ------------------------------------
        # Header Formatting
        # ------------------------------------

        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # ------------------------------------
        # Freeze Header
        # ------------------------------------

        ws.freeze_panes = "A2"

        # ------------------------------------
        # Auto Width
        # ------------------------------------

        for column in ws.columns:

            length = max(

                len(str(cell.value))

                if cell.value is not None

                else 0

                for cell in column

            )

            ws.column_dimensions[
                column[0].column_letter
            ].width = length + 3

        # ------------------------------------
        # Percentile Formatting
        # ------------------------------------

        headers = {}

        for cell in ws[1]:

            headers[cell.value] = cell.column

        percentile_cols = [

            col

            for name, col in headers.items()

            if name is not None

            and "percentile" in str(name).lower()

        ]

        for col in percentile_cols:

            for row in range(2, ws.max_row + 1):

                value = ws.cell(
                    row,
                    col
                ).value

                if value is None:

                    continue

                if value >= 75:

                    ws.cell(
                        row,
                        col
                    ).fill = green

                elif value >= 25:

                    ws.cell(
                        row,
                        col
                    ).fill = yellow

                else:

                    ws.cell(
                        row,
                        col
                    ).fill = red

    wb.save(OUTPUT)

    print("\nWorkbook Formatting Completed.")
# ---------------------------------------------------------
# Final Workbook Formatting
# ---------------------------------------------------------

def finalize_workbook():

    wb = load_workbook(OUTPUT)

    gold = PatternFill(
        start_color="FFD966",
        end_color="FFD966",
        fill_type="solid"
    )

    for ws in wb.worksheets:

        headers = {}

        for cell in ws[1]:
            headers[cell.value] = cell.column

        score_col = headers.get("composite_quality_score")

        if score_col:

            highest = -999999
            benchmark_row = 2

            for row in range(2, ws.max_row + 1):

                value = ws.cell(row, score_col).value

                if value is None:
                    continue

                if value > highest:

                    highest = value
                    benchmark_row = row

            # Highlight benchmark company

            for cell in ws[benchmark_row]:

                cell.fill = gold

            # Median Row

            values = []

            for row in range(2, ws.max_row + 1):

                value = ws.cell(row, score_col).value

                if isinstance(value, (int, float)):
                    values.append(value)

            if len(values) > 0:

                median = round(pd.Series(values).median(), 2)

                last = ws.max_row + 2

                ws.cell(last, 1).value = "Peer Group Median"

                ws.cell(last, score_col).value = median

                ws.cell(last, 1).font = Font(bold=True)

                ws.cell(last, score_col).font = Font(bold=True)

    wb.save(OUTPUT)

    print("\nWorkbook Final Formatting Completed.")

# ---------------------------------------------------------
# Main
# ---------------------------------------------------------

def main():

    preview()

    print("\nGenerating Excel Report...\n")

    save_workbook()

    print("\nDay 20  Completed Successfully.")


# ---------------------------------------------------------
# Execute
# ---------------------------------------------------------

if __name__ == "__main__":

    main()

# ---------------------------------------------------------
# Preview
# ---------------------------------------------------------

def preview():

    df = prepare_data()

    print("=" * 60)
    print("SPRINT 3 - DAY 20")
    print("PEER COMPARISON REPORT")
    print("=" * 60)

    print()

    print("Companies Loaded :", len(df))

    print()

    print(
        df[
            [
                "company_id",
                "company_name",
                "peer_group_name"
            ]
        ].head(20)
    )


if __name__ == "__main__":

    preview()