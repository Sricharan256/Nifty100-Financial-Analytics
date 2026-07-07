"""
peer_export.py

Sprint 3 - Day 18

Generate Peer Comparison Excel Report
"""

import sqlite3
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

DB_PATH = "db/nifty100.db"

OUTPUT = "output/peer_comparison.xlsx"


# ---------------------------------------------------------
# Load Peer Percentiles
# ---------------------------------------------------------

def load_peer_percentiles():

    conn = sqlite3.connect(DB_PATH)

    df = pd.read_sql(
        """
        SELECT *
        FROM peer_percentiles
        """,
        conn
    )

    conn.close()

    return df


# ---------------------------------------------------------
# Load Company Information
# ---------------------------------------------------------

def load_companies():

    conn = sqlite3.connect(DB_PATH)

    companies = pd.read_sql(
        """
        SELECT
            id,
            company_name
        FROM companies
        """,
        conn
    )

    conn.close()

    companies.rename(
        columns={
            "id": "company_id"
        },
        inplace=True
    )

    return companies


# ---------------------------------------------------------
# Merge Data
# ---------------------------------------------------------

def prepare_data():

    peer_df = load_peer_percentiles()

    companies = load_companies()

    df = pd.merge(

        peer_df,

        companies,

        on="company_id",

        how="left"

    )

    return df


# ---------------------------------------------------------
# Get Peer Groups
# ---------------------------------------------------------

def get_peer_groups(df):

    groups = sorted(

        df["peer_group_name"]

        .dropna()

        .unique()

    )

    return groups


# ---------------------------------------------------------
# Prepare Workbook Data
# ---------------------------------------------------------

def prepare_sheet(df, peer_group):

    sheet = df[
        df["peer_group_name"] == peer_group
    ].copy()

    sheet = sheet.sort_values(

        by="percentile_rank",

        ascending=False

    )

    columns = [

        "company_id",

        "company_name",

        "metric",

        "value",

        "percentile_rank",

        "year"

    ]

    return sheet[columns]
# ---------------------------------------------------------
# Export Excel
# ---------------------------------------------------------

def export_excel():

    os.makedirs("output", exist_ok=True)

    df = prepare_data()

    peer_groups = get_peer_groups(df)

    with pd.ExcelWriter(
        OUTPUT,
        engine="openpyxl"
    ) as writer:

        for group in peer_groups:

            sheet = prepare_sheet(df, group)

            # Excel sheet names cannot exceed 31 characters
            sheet_name = group[:31]

            sheet.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

    print("=" * 60)
    print("SPRINT 3 - DAY 18")
    print("PEER COMPARISON REPORT")
    print("=" * 60)

    print(f"Peer Groups Exported : {len(peer_groups)}")

    format_excel()

    print("\nWorkbook formatted successfully.") 


# ---------------------------------------------------------
# Preview
# ---------------------------------------------------------

def preview():

    df = prepare_data()

    print("\nPreview\n")

    print(
        df[
            [
                "company_id",
                "company_name",
                "peer_group_name",
                "metric",
                "percentile_rank"
            ]
        ].head(20)
    )

# ---------------------------------------------------------
# Format Excel Workbook
# ---------------------------------------------------------
def format_excel():

    wb = load_workbook(OUTPUT)

    green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    yellow = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    gold = PatternFill(
        start_color="FFD966",
        end_color="FFD966",
        fill_type="solid"
    )

    for ws in wb.worksheets:

        # -------------------------
        # Header Formatting
        # -------------------------

        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # -------------------------
        # Freeze Header
        # -------------------------

        ws.freeze_panes = "A2"

        # -------------------------
        # Auto Width
        # -------------------------

        for column in ws.columns:

            length = max(

                len(str(cell.value))

                if cell.value is not None

                else 0

                for cell in column

            )

            ws.column_dimensions[
                column[0].column_letter
            ].width = length + 4

        # -------------------------
        # Conditional Formatting
        # -------------------------

        headers = {}

        for cell in ws[1]:

            headers[cell.column] = cell.value

        percentile_col = None

        company_col = None

        for col, name in headers.items():

            if name == "percentile_rank":
                percentile_col = col

            if name == "company_name":
                company_col = col

        if percentile_col:

            for row in range(2, ws.max_row + 1):

                value = ws.cell(
                    row,
                    percentile_col
                ).value

                if value is None:
                    continue

                if value >= 75:

                    ws.cell(
                        row,
                        percentile_col
                    ).fill = green

                elif value >= 25:

                    ws.cell(
                        row,
                        percentile_col
                    ).fill = yellow

                else:

                    ws.cell(
                        row,
                        percentile_col
                    ).fill = red

        # -------------------------
        # Highlight Benchmark Company
        # (Highest Percentile)
        # -------------------------

        if percentile_col:

            highest = -1
            benchmark_row = None

            for row in range(2, ws.max_row + 1):

                value = ws.cell(
                    row,
                    percentile_col
                ).value

                if value is not None and value > highest:

                    highest = value
                    benchmark_row = row

            if benchmark_row:

                for cell in ws[benchmark_row]:

                    cell.fill = gold

        # -------------------------
        # Median Summary Row
        # -------------------------

        if percentile_col:

            values = []

            for row in range(2, ws.max_row + 1):

                value = ws.cell(
                    row,
                    percentile_col
                ).value

                if value is not None:

                    values.append(value)

            if values:

                median = round(
                    sum(values) / len(values),
                    2
                )

                last = ws.max_row + 2

                ws.cell(
                    last,
                    1
                ).value = "Median Percentile"

                ws.cell(
                    last,
                    percentile_col
                ).value = median

                ws.cell(
                    last,
                    1
                ).font = Font(bold=True)

    wb.save(OUTPUT)


# ---------------------------------------------------------
# Main
# ---------------------------------------------------------
if __name__ == "__main__":
    
    preview()

    export_excel()

    print("\nDay 18 Export Completed Successfully.")

